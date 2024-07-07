import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import smtplib
import email.message
from datetime import date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage
from openpyxl import workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles import NamedStyle

# Encontrar o ano atual.
ano_atual = date.today().year

# Ler e organizar os dados da planilha
df = pd.read_excel('TabeladeVendas.xlsm', sheet_name='Planilha3')
df['Data'] = pd.to_datetime(df['Data'])
df['Mês'] = df['Data'].dt.to_period('M')
df['Ano'] = df['Data'].dt.to_period('Y')
df_ano_atual = df.loc[(df['Ano'] == str(ano_atual))]

# Criar a base do relatório
vendas_por_mes = df_ano_atual.groupby('Mês')['Valor Vendido'].sum().reset_index()
vendas_por_mes.to_excel('relatorio_mensal.xlsx', index=False, sheet_name='Relatório Mensal')
print(f"Relatório mensal salvo em: 'relatorio_mensal.xlsx'")

vendas_por_ano = df.groupby('Ano')['Valor Vendido'].sum().reset_index()
vendas_por_ano.to_excel('relatorio_anual.xlsx', index=False, sheet_name='Relatório Anual')
print(f"Relatório anual salvo em: 'relatorio_anual.xlsx'")

# Estilizar os relatórios
def tratar_planilhas(arquivo):
    # Busca os dados dos funcionários
    func = pd.read_excel('funcionarios-da-empresa.xlsx')

    # Define o custo
    if arquivo == 'relatorio_mensal.xlsx':
        custo = func['Salário (R$)'].sum() + 10000
    elif arquivo == 'relatorio_anual.xlsx':
        custo = (func['Salário (R$)'].sum() + 10000) * 12

    # Define as cores e outras propiedades
    verde = PatternFill(start_color='FF63B005', end_color='FF63B005', fill_type='solid')
    vermelho = PatternFill(start_color='FFE02F02', end_color='FFE02F02', fill_type='solid')
    amarelo = PatternFill(start_color='FFFFDA00', end_color='FFFFDA00', fill_type='solid')

    preto = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    fonte_branca = Font(color='FFFFFF')

    centralizado = Alignment(horizontal='center')

    # Abre e edita a planilha
    planilha = load_workbook(arquivo)
    aba_ativa = planilha.active
    
    for celula in aba_ativa['B'][1:]:
        linha = celula.row
        if celula.value > custo: 
            aba_ativa[f'B{linha}'].fill = verde
            aba_ativa[f'B{linha}'].number_format = 'R$ #,##0.00'
        elif celula.value == custo: 
            aba_ativa[f'B{linha}'].fill = amarelo
            aba_ativa[f'B{linha}'].number_format = 'R$ #,##0.00'
        elif celula.value < custo: 
            aba_ativa[f'B{linha}'].fill = vermelho
            aba_ativa[f'B{linha}'].number_format = 'R$ #,##0.00'
    
    aba_ativa.column_dimensions['A'].width = 15
    aba_ativa.column_dimensions['B'].width = 35
    aba_ativa.column_dimensions['D'].width = 45

    # Cria a legenda
    aba_ativa['D1'].value = 'Legenda'
    aba_ativa['D1'].fill = preto
    aba_ativa['D1'].font = fonte_branca
    aba_ativa['D1'].alignment = centralizado
    
    aba_ativa['D2'].value = 'Faturamento > Custo'
    aba_ativa['D2'].fill = verde
    aba_ativa['D2'].alignment = centralizado
    
    aba_ativa['D3'].value = 'Faturamento = Custo'
    aba_ativa['D3'].fill = amarelo
    aba_ativa['D3'].alignment = centralizado
    
    aba_ativa['D4'].value = 'Faturamento < Custo'
    aba_ativa['D4'].fill = vermelho
    aba_ativa['D4'].alignment = centralizado

    planilha.save(arquivo)
        
tratar_planilhas('relatorio_mensal.xlsx')
tratar_planilhas('relatorio_anual.xlsx')

# Separar os dados relevantes para o email
mes_menos_vendas = vendas_por_mes.loc[vendas_por_mes['Valor Vendido'].idxmin()] # Linha completa
mes_mais_vendas = vendas_por_mes.loc[vendas_por_mes['Valor Vendido'].idxmax()] # Linha completa

mmaisv = str(mes_mais_vendas['Mês'])
mmenosv = str(mes_menos_vendas['Mês'])

def traducao_de_mes(mes):
    # Trauz o nome do mês para pt-br
    if mes[5:7] == '01':
        mes = 'Janeiro'
    elif mes[5:7] == '02':
        mes = 'Fevereiro'
    elif mes[5:7] == '03':
        mes = 'Março'
    elif mes[5:7] == '04':
        mes = 'Abril'
    elif mes[5:7] == '05':
        mes = 'Maio'
    elif mes[5:7] == '06':
        mes = 'Junho'
    elif mes[5:7] == '07':
        mes = 'Julho'
    elif mes[5:7] == '08':
        mes = 'Agosto'
    elif mes[5:7] == '09':
        mes = 'Setembro'
    elif mes[5:7] == '10':
        mes = 'Outubro'
    elif mes[5:7] == '11':
        mes = 'Novembro'
    else: 
        mes = 'Dezembro'
    print('Tradução dos meses realizada!')
    return mes


def gerar_graficos():
    # Gerar uma figura
    fig, axs = plt.subplots(1, 2, figsize=(14, 6))

    # Gera um gráfico em barra.
    axs[0].bar(vendas_por_mes['Mês'].astype(str), vendas_por_mes['Valor Vendido'], label='Vendas realizadas')
    axs[0].set_xlabel('Mês')
    axs[0].set_ylabel('Valor Vendido')
    axs[0].set_title(f'Vendas por Mês no Ano de {str(ano_atual)}')
    axs[0].legend()

    # Gera um gráfico em linha.
    axs[1].plot(vendas_por_ano['Ano'].astype(str), vendas_por_ano['Valor Vendido'], label='Vendas por Ano', color='orange')
    axs[1].set_xlabel('Ano')
    axs[1].set_ylabel('Valor Vendido')
    axs[1].set_title('Vendas por Ano')
    axs[1].legend()

    plt.subplots_adjust(wspace=0.8)
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    plt.savefig('graficos.png')
    print('Gráfico gerado!')

gerar_graficos()

# Preparar a mensagem do email
AnoAtual_MesMaisVendas_traduzido = traducao_de_mes(mmaisv)
AnoAtual_MesMenosVendas_traduzido = traducao_de_mes(mmenosv)

assunto_do_email = f'Relatório de Vendas - {ano_atual}'
corpo_do_email = f"""
<h1>Relatório de Vendas - {ano_atual}</h1>
<p>Prezado(a),</p>
<p>Espero que esteja bem. Gostaria de compartilhar com você o relatório de vendas por mês referente ao ano de {ano_atual} e um histórico de faturamento anual ao longo dos anos da empresa.</p>
<h2>Resumo de Vendas:</h2>
<p>No mês de {AnoAtual_MesMaisVendas_traduzido} alcançamos a maior receita registrada no ano, destacando o desempenho excepcional da nossa equipe. Por outro lado, no mês de {AnoAtual_MesMenosVendas_traduzido}, observamos o desafio de ajustar nossa abordagem de mercado, a fim de elevar o faturamento.</p>
<img src="cid:graficos">
<h2>Anexos:</h2>
<p>Como mencionado anteriormente, seguem anexas duas planilhas importantes para análise detalhada: </p>
<ol>
    <li>relatorio_mensal: Consta as vendas por mês no ano de {ano_atual}</li>
    <li>relatorio_anual: Consta o histórico de faturamento da empresa por ano</li>
</ol>
<h2>Considerações Finais:</h2>
<p>Agradecemos a todos pelo empenho e dedicação ao longo deste ano. Estamos confiantes de que, com as estratégias adequadas, continuaremos a crescer para atingir nossos objetivos de vendas e assegurar o desenvolvimento sustentável da empresa.</p>
<p>Atenciosamente,</p>
<p>Henrique Guimarães</p>
<p>CEO</p>
"""

def localizar_emails(planilha_dos_funcionarios):
    emails = []
    funcionarios = pd.read_excel(planilha_dos_funcionarios)
    # Coleta os emails dos funcionários
    for index, linha in funcionarios.iterrows():
        if linha['Cargo'] in ['Vendedor', 'Gerente', 'Contador','Vendedora','Contadora']:
            emails.append(linha['E-mail'])
    print('E-mails salvos!')
    return emails

destino = localizar_emails('funcionarios-da-empresa.xlsx')

def enviar_email(assunto,destinatarios,corpo_email):
    # Logar no email
    login = 'login'
    senha = 'senha'

    for email in range(len(destinatarios)):
        # Criar o servidor de email
        msg = MIMEMultipart()
        msg['Subject'] = assunto
        msg['From'] = login
        msg['To'] = destinatarios[email]

        # Adicionar o corpo do email
        msg.attach(MIMEText(corpo_email, 'html'))

        # Permite a inserção da imagem
        with open('graficos.png', 'rb') as arquivo_imagem:
            imagem = MIMEImage(arquivo_imagem.read())
            imagem.add_header('Content-ID', '<graficos>')
            imagem.add_header('Content-Disposition', 'inline', filename='graficos.png')
            msg.attach(imagem)

        # Prepara os anexos
        anexo_mensal = 'relatorio_mensal.xlsx'
        anexo_anual = 'relatorio_anual.xlsx'
        with open(anexo_mensal, 'rb') as attachment:
            part = MIMEApplication(attachment.read(), Name=anexo_mensal)
            part['Content-Disposition'] = f'attachment; filename="{anexo_mensal}"'
            msg.attach(part)
        with open(anexo_anual, 'rb') as attachment:
            part = MIMEApplication(attachment.read(), Name=anexo_anual)
            part['Content-Disposition'] = f'attachment; filename="{anexo_anual}"'
            msg.attach(part)

        # Enviar email
        server = smtplib.SMTP('smtp.gmail.com: 587')
        server.starttls()
        server.login(login, senha)
        server.sendmail(login, destinatarios[email], msg.as_string().encode('utf-8'))
        server.quit()
        print(f'Email enviado para {destinatarios[email]}!')

enviar_email(assunto_do_email,destino,corpo_do_email)